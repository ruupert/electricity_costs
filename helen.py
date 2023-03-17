import os
import time
import datetime
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.firefox.options import Options

class Helen:

    def __init_geckodriver(self):
        options = Options()
        options.add_argument('-headless')
        options.add_argument("-profile")
        options.add_argument(".selenium")
        #options.set_preference("browser.download.folderList", 2)
        #options.set_preference("browser.download.dir", './dl')

        return webdriver.Firefox(options=options,
                                service_args=["--marionette-port", "2828"]
        )

    def __init__(self, database, username, password):

        self.start_date = datetime.now() + timedelta(days=-7)
        self.end_date = datetime.now() + timedelta(days=-1)

        self.driver = self.__init_geckodriver()

        try:

            self.__login(username, password)
            self.__get_consumption()

            if self.fpath != None:
                wb_obj = openpyxl.load_workbook(Path(self.fpath))
                sheet = wb_obj.active
                for row in sheet.iter_rows():
                    if row[0].value != "Ajankohta":
                        kw = row[1].value
                        if kw == None:
                            kw = 0.0
                        data_tuple = (datetime.strftime(row[0].value, "%Y-%m-%d %H:00:00"),kw)
                        database.insert_or_update('kwh', data_tuple)

                wb_obj.close()
                os.remove(self.fpath)
        except:
            None
        self.driver.close()


    def __login(self, username, password):
        self.driver.get("https://www.helen.fi/kirjautuminen")
        self.driver.switch_to.frame(0)
        self.driver.find_element(By.NAME, 'username').send_keys(username)
        self.driver.find_element(By.NAME, 'password').send_keys(password)
        self.driver.find_element(By.XPATH, '/html/body/table/tbody/tr/td/div/div[2]/div[2]/div/div[3]/div/form/div/span/input').click()
        time.sleep(10) # for now wait here statically the redirections to finish
        self.driver.switch_to.default_content()
        #WebDriverWait(self.driver, 3).until(EC.presence_of_element_located((By.CLASS_NAME, "recharts-wrapper")))

    def __get_consumption(self):
        self.driver.get("https://web.omahelen.fi/personal/reports/electricity-consumption")
        WebDriverWait(self.driver, 3).until(EC.presence_of_element_located((By.CLASS_NAME, "recharts-surface")))

        for elem in self.driver.find_elements(By.XPATH, "//button"):
            if elem.get_attribute("data-cy") == "downloadExcel":
                elem.click()

        self.driver.find_element(By.XPATH, '//*[@id="downshift-0-toggle-button"]').click()
        self.driver.find_element(By.XPATH, '//*[@id="downshift-0-item-3"]').click()
        self.driver.find_element(By.XPATH, '//*[@id="date1"]').send_keys(Keys.CONTROL, "a", Keys.DELETE)
        self.driver.find_element(By.XPATH, '//*[@id="date1"]').send_keys(self.start_date.strftime("%d.%m.%Y"))
        self.driver.find_element(By.XPATH, '//*[@id="date2"]').send_keys(Keys.CONTROL, "a", Keys.DELETE)
        self.driver.find_element(By.XPATH, '//*[@id="date2"]').send_keys(self.end_date.strftime("%d.%m.%Y"))
        self.driver.find_element(By.XPATH, '//*[@id="downshift-1-toggle-button"]').click()
        self.driver.find_element(By.XPATH, '//*[@id="downshift-1-item-0"]').click()
        self.driver.find_element(By.XPATH, '/html/body/div[2]/div/div[3]/div/button[2]').click()
        self.fpath = f"./dl/electricity_report_{ self.start_date.strftime('%d_%m_%Y') }_{ self.end_date.strftime('%d_%m_%Y') }.xlsx"
        
        # wait for the file to be downloaded
        for i in range(1,5):
            time.sleep(5)
            if os.path.exists(self.fpath):
                break
        
  




        ## here would be downloading the electricity eur per hour report but helen exporter returns the same as above so no exportable data
        #self.driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/main/div/div[2]/div/div[1]/div[2]/div/div[2]/div/div/div/button[2]').click()
        #time.sleep(5)
        #self.driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/main/div/div[2]/div/div[1]/div[1]/div[1]/button').click()
        #self.driver.find_element(By.XPATH, '//*[@id="downshift-28-toggle-button"]').click()
        #self.driver.find_element(By.XPATH, '//*[@id="downshift-28-item-3"]').click()
        #self.driver.find_element(By.XPATH, '//*[@id="downshift-29-toggle-button"]').click()
        #self.driver.find_element(By.XPATH, '//*[@id="downshift-29-item-0"]').click()
        #self.driver.find_element(By.XPATH, '//*[@id="date1"]').send_keys(Keys.CONTROL, "a", Keys.DELETE)
        #self.driver.find_element(By.XPATH, '//*[@id="date1"]').send_keys(start_date.strftime("%d.%m.%Y"))
        #self.driver.find_element(By.XPATH, '//*[@id="date2"]').send_keys(Keys.CONTROL, "a", Keys.DELETE)
        #self.driver.find_element(By.XPATH, '//*[@id="date2"]').send_keys(end_date.strftime("%d.%m.%Y"))
        #self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div[3]/div/button[2]').click()
        #time.sleep(7)
        #
        #wb_obj = openpyxl.load_workbook(Path(fpath))
        #sheet = wb_obj.active
        #for row in sheet.iter_rows():
        #    if row[0].value != "Ajankohta":
        #        for cell in row: 
        #            print(cell.value, end=" ")
        #        print("")
        #sheet = None
        #wb_obj.close()
        #os.remove(fpath)